"""
Section management endpoints for SmartAttend.
CRUD for academic sections (class groups), with bulk CSV/Excel upload support.
"""

from fastapi import APIRouter, Depends, HTTPException, UploadFile, File
from sqlalchemy.orm import Session
from typing import List, Optional
from ..core.database import get_db
from ..models.models import Section, User
from ..api import deps
from pydantic import BaseModel
import io

router = APIRouter()

class SectionCreate(BaseModel):
    name: str
    academic_year: str
    department: Optional[str] = None
    year: Optional[int] = None
    semester: Optional[int] = None

class SectionUpdate(BaseModel):
    name: str
    academic_year: str
    department: Optional[str] = None
    year: Optional[int] = None
    semester: Optional[int] = None

class SectionResponse(BaseModel):
    id: int
    name: str
    academic_year: str
    department: Optional[str] = None
    year: Optional[int] = None
    semester: Optional[int] = None

@router.post("/", response_model=SectionResponse)
def create_section(
    section: SectionCreate,
    db: Session = Depends(get_db),
    current_admin: User = Depends(deps.get_current_active_admin)
):
    """Create a new section (Admin only)"""
    existing = db.query(Section).filter(
        Section.name == section.name,
        Section.department == section.department,
        Section.year == section.year,
        Section.academic_year == section.academic_year
    ).first()
    if existing:
        raise HTTPException(status_code=400, detail=f"Section '{section.name}' already exists for {section.department} Year {section.year} ({section.academic_year})")

    db_section = Section(
        name=section.name,
        academic_year=section.academic_year,
        department=section.department,
        year=section.year,
        semester=section.semester
    )
    db.add(db_section)
    db.commit()
    db.refresh(db_section)
    return SectionResponse(id=db_section.id, name=db_section.name, academic_year=db_section.academic_year, department=db_section.department, year=db_section.year, semester=db_section.semester)

@router.get("/", response_model=List[SectionResponse])
def list_sections(
    skip: int = 0,
    limit: int = 10000,
    db: Session = Depends(get_db),
    current_user: User = Depends(deps.get_current_user)
):
    """List all sections"""
    sections = db.query(Section).offset(skip).limit(limit).all()
    return [SectionResponse(id=s.id, name=s.name, academic_year=s.academic_year, department=s.department, year=s.year, semester=s.semester) for s in sections]

@router.get("/{section_id}", response_model=SectionResponse)
def get_section(
    section_id: int,
    db: Session = Depends(get_db),
    current_user: User = Depends(deps.get_current_user)
):
    """Get section by ID"""
    section = db.query(Section).filter(Section.id == section_id).first()
    if not section:
        raise HTTPException(status_code=404, detail="Section not found")
    return SectionResponse(id=section.id, name=section.name, academic_year=section.academic_year, department=section.department, year=section.year, semester=section.semester)

@router.put("/{section_id}", response_model=SectionResponse)
def update_section(
    section_id: int,
    section_update: SectionUpdate,
    db: Session = Depends(get_db),
    current_admin: User = Depends(deps.get_current_active_admin)
):
    """Update section (Admin only)"""
    section = db.query(Section).filter(Section.id == section_id).first()
    if not section:
        raise HTTPException(status_code=404, detail="Section not found")

    existing = db.query(Section).filter(
        Section.name == section_update.name,
        Section.department == section_update.department,
        Section.year == section_update.year,
        Section.academic_year == section_update.academic_year,
        Section.id != section_id
    ).first()
    if existing:
        raise HTTPException(status_code=400, detail=f"Section '{section_update.name}' already exists for {section_update.department} Year {section_update.year} ({section_update.academic_year})")

    section.name = section_update.name
    section.academic_year = section_update.academic_year
    section.department = section_update.department
    section.year = section_update.year
    section.semester = section_update.semester
    db.commit()
    db.refresh(section)
    return SectionResponse(id=section.id, name=section.name, academic_year=section.academic_year, department=section.department, year=section.year, semester=section.semester)

@router.delete("/{section_id}")
def delete_section(
    section_id: int,
    db: Session = Depends(get_db),
    current_admin: User = Depends(deps.get_current_active_admin)
):
    """Delete section (Admin only)"""
    section = db.query(Section).filter(Section.id == section_id).first()
    if not section:
        raise HTTPException(status_code=404, detail="Section not found")

    if section.students:
        raise HTTPException(
            status_code=400,
            detail=f"Cannot delete section with {len(section.students)} students. Reassign students first."
        )

    db.delete(section)
    db.commit()
    return {"message": "Section deleted successfully"}

@router.post("/bulk-upload")
def bulk_upload_sections(
    file: UploadFile = File(...),
    db: Session = Depends(get_db),
    current_admin: User = Depends(deps.get_current_active_admin)
):
    """Bulk upload sections from CSV or Excel file"""
    filename = file.filename.lower()
    content = file.file.read()

    if filename.endswith('.csv'):
        import csv
        reader = csv.DictReader(io.StringIO(content.decode('utf-8')))
        rows = list(reader)
    elif filename.endswith(('.xlsx', '.xls')):
        import openpyxl
        wb = openpyxl.load_workbook(io.BytesIO(content))
        ws = wb.active
        headers = [cell.value.strip().lower() if cell.value else '' for cell in ws[1]]
        rows = []
        for row in ws.iter_rows(min_row=2, values_only=True):
            row_dict = {headers[i]: (str(row[i]).strip() if row[i] is not None else '') for i in range(len(headers))}
            rows.append(row_dict)
    else:
        raise HTTPException(status_code=400, detail="Only CSV and Excel files are supported")

    created = []
    skipped = []
    errors = []

    for i, row in enumerate(rows, start=2):
        name = row.get('name', '').strip()
        academic_year = row.get('academic_year', '').strip()
        department = row.get('department', '').strip()
        year_val = row.get('year', '').strip()
        semester_val = row.get('semester', '').strip()

        if not name:
            errors.append(f"Row {i}: Missing section name")
            continue

        existing = db.query(Section).filter(
            Section.name == name,
            Section.department == (department or None),
            Section.year == (int(year_val) if year_val else None),
            Section.academic_year == (academic_year or None)
        ).first()
        if existing:
            skipped.append(f"{department}-{name} (already exists)")
            continue

        try:
            year_int = int(year_val) if year_val else None
        except ValueError:
            year_int = None

        try:
            semester_int = int(semester_val) if semester_val else None
        except ValueError:
            semester_int = None

        db_section = Section(
            name=name,
            academic_year=academic_year or None,
            department=department or None,
            year=year_int,
            semester=semester_int
        )
        db.add(db_section)
        created.append(name)

    db.commit()

    return {
        "message": f"Bulk upload complete: {len(created)} created, {len(skipped)} skipped, {len(errors)} errors",
        "created": created,
        "skipped": skipped,
        "errors": errors
    }
